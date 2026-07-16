# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable


SOURCE_DIR = Path("data_sources")
EXPORT_DIR = SOURCE_DIR / "competitor_sheet_exports"
MANUAL_CSV = SOURCE_DIR / "observed_archive_supplement.csv"
OUTPUT_CSV = SOURCE_DIR / "observed_archive_supplement.csv"

VIDEO_ID_RE = re.compile(r"(?:v=|youtu\.be/|shorts/)([A-Za-z0-9_-]{11})")
FIELDS = [
    "video_id",
    "channel_id",
    "channel_title",
    "video_title",
    "published_at",
    "observed_view_count",
    "observed_like_count",
    "observed_comment_count",
    "observed_at",
    "source_name",
    "archive_type",
    "thumbnail_url",
    "thumbnail_gcs_uri",
    "thumbnail_saved_url",
]


def main() -> int:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    rows = []
    rows.extend(read_existing_manual_rows())
    rows.extend(read_export_rows())
    rows = dedupe_rows(rows)
    write_rows(OUTPUT_CSV, rows)
    print(f"wrote {len(rows)} rows to {OUTPUT_CSV}")
    return 0


def read_existing_manual_rows() -> list[dict[str, str]]:
    if not MANUAL_CSV.exists():
        return []
    with MANUAL_CSV.open("r", newline="", encoding="utf-8-sig") as f:
        return [normalize_output_row(row) for row in csv.DictReader(f) if row.get("video_id")]


def read_export_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in sorted(EXPORT_DIR.rglob("*")):
        if path.suffix.lower() == ".csv":
            rows.extend(read_csv_export(path))
        elif path.suffix.lower() == ".xlsx":
            rows.extend(read_xlsx_export(path))
    return rows


def read_csv_export(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        raw_rows = list(csv.reader(f))
    return normalize_table(raw_rows, source_name=path.stem)


def read_xlsx_export(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required for .xlsx imports. Run with: uv run --with openpyxl python build_observed_archive_supplement.py") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    for ws in wb.worksheets:
        raw_rows = [[cell_to_text(cell.value) for cell in row] for row in ws.iter_rows()]
        rows.extend(normalize_table(raw_rows, source_name=f"{path.stem}:{ws.title}"))
    return rows


def normalize_table(raw_rows: list[list[str]], source_name: str) -> list[dict[str, str]]:
    header_idx = find_header_row(raw_rows)
    if header_idx is None:
        return []
    headers = [compact(cell) for cell in raw_rows[header_idx]]
    indexes = build_indexes(headers)
    required = ["video_url", "video_title", "view_count"]
    if any(key not in indexes for key in required):
        return []

    rows = []
    for raw in raw_rows[header_idx + 1 :]:
        row = row_from_indexes(raw, indexes)
        video_id = extract_video_id(row.get("video_url", ""))
        if not video_id:
            continue
        observed_at = row.get("observed_at") or infer_observed_at(source_name)
        channel_title = row.get("channel_title", "") or infer_channel_title(source_name)
        rows.append(
            {
                "video_id": video_id,
                "channel_id": row.get("channel_id", ""),
                "channel_title": channel_title,
                "video_title": row.get("video_title", ""),
                "published_at": normalize_date(row.get("published_at", "")),
                "observed_view_count": digits(row.get("view_count", "")),
                "observed_like_count": digits(row.get("like_count", "")),
                "observed_comment_count": digits(row.get("comment_count", "")),
                "observed_at": normalize_date(observed_at),
                "source_name": source_name,
                "archive_type": "competitor_sheet_archive",
                "thumbnail_url": f"https://i.ytimg.com/vi/{video_id}/maxresdefault.jpg",
                "thumbnail_gcs_uri": "",
                "thumbnail_saved_url": "",
            }
        )
    return rows


def find_header_row(rows: list[list[str]]) -> int | None:
    for i, row in enumerate(rows[:10]):
        cells = [compact(cell) for cell in row]
        if ("動画URL" in cells or "URL" in cells) and ("動画タイトル" in cells or "タイトル" in cells) and ("視聴回数" in cells or "再生回数" in cells):
            return i
    return None


def build_indexes(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, header in enumerate(headers):
        if header == "チャンネルID":
            mapping.setdefault("channel_id", i)
        elif header in {"動画URL", "URL"}:
            mapping.setdefault("video_url", i)
        elif header in {"動画タイトル", "タイトル"}:
            mapping.setdefault("video_title", i)
        elif header == "チャンネル名":
            mapping.setdefault("channel_title", i)
        elif header in {"動画公開日", "投稿日"}:
            mapping.setdefault("published_at", i)
        elif header in {"視聴回数", "再生回数"}:
            mapping.setdefault("view_count", i)
        elif header == "コメント数":
            mapping.setdefault("comment_count", i)
        elif header == "高評価数":
            mapping.setdefault("like_count", i)
        elif header == "リサーチ日時":
            mapping.setdefault("observed_at", i)
    return mapping


def row_from_indexes(raw: list[str], indexes: dict[str, int]) -> dict[str, str]:
    row = {}
    for key, idx in indexes.items():
        row[key] = compact(raw[idx]) if idx < len(raw) else ""
    return row


def dedupe_rows(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    by_key: dict[tuple[str, str, str], dict[str, str]] = {}
    for row in rows:
        normalized = normalize_output_row(row)
        key = (normalized["video_id"], normalized["observed_at"], normalized["source_name"])
        # The workbook exports are read after the previously generated CSV.
        # Prefer the latest source row so a corrected import can replace stale
        # values that were written by an older version of this script.
        by_key[key] = normalized
    return list(by_key.values())


def normalize_output_row(row: dict[str, object]) -> dict[str, str]:
    normalized = {field: compact(row.get(field, "")) for field in FIELDS}
    if normalized["video_id"] and not normalized["thumbnail_url"]:
        normalized["thumbnail_url"] = f"https://i.ytimg.com/vi/{normalized['video_id']}/maxresdefault.jpg"
    if not normalized["archive_type"]:
        normalized["archive_type"] = "competitor_sheet_archive"
    return normalized


def write_rows(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def extract_video_id(url: str) -> str:
    match = VIDEO_ID_RE.search(url)
    return match.group(1) if match else ""


def infer_observed_at(source_name: str) -> str:
    match = re.search(r"20\d{6}", source_name)
    if not match:
        return ""
    text = match.group(0)
    return f"{text[:4]}-{text[4:6]}-{text[6:8]}"


def infer_channel_title(source_name: str) -> str:
    sheet_name = source_name.split(":", 1)[-1].strip()
    if sheet_name.startswith("動画_"):
        return sheet_name.replace("動画_", "", 1).strip()
    return ""


def normalize_date(value: str) -> str:
    text = compact(value)
    match = re.search(r"(20\d{2})[/-](\d{1,2})[/-](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return text


def digits(value: str) -> str:
    text = compact(value).replace(",", "").replace("，", "")
    if not text:
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return ""
    try:
        number = Decimal(match.group(0))
    except InvalidOperation:
        return ""
    if number < 0:
        return ""
    return str(int(number))


def compact(value: object) -> str:
    return " ".join(str(value or "").replace("\r", "\n").split())


def cell_to_text(value: object) -> str:
    return "" if value is None else str(value)


if __name__ == "__main__":
    raise SystemExit(main())
